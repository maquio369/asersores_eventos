from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth import update_session_auth_hash
from django.http import HttpResponse, Http404, JsonResponse, HttpResponsePermanentRedirect
from django.utils import timezone
from datetime import timedelta, datetime
from .models import Evento, Municipio, CustomUser, Dependencia
from .forms import EventoForm, FiltroEventosForm, PerfilUsuarioForm, CustomPasswordChangeForm, AdminCrearUsuarioForm, AdminEditarUsuarioForm
from .decorators import captura_or_admin_required, admin_required, active_user_required
from django.core.management import call_command
from io import StringIO
import calendar
import json
from collections import defaultdict
import os
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def redirect_to_http(request):
    """Redireccionar peticiones HTTPS a HTTP"""
    if request.is_secure():
        url = request.build_absolute_uri().replace('https://', 'http://')
        return HttpResponsePermanentRedirect(url)
    return None

def enviar_notificacion_evento(evento, subject, tipo_accion='creado'):
    """
    Función auxiliar para enviar notificaciones por correo electrónico sobre un evento.
    :param tipo_accion: 'creado' o 'actualizado' para cambiar el texto en el correo.
    """
    # No enviar correo si el usuario no tiene uno configurado
    if not evento.usuario_creador.email:
        return

    contexto_email = {
        'evento': evento,
        'subject': subject,
        'tipo_accion': tipo_accion,
    }
    
    html_message = render_to_string('emails/notificacion_evento.html', contexto_email)
    
    destinatarios = [evento.usuario_creador.email]
    if hasattr(settings, 'ADMIN_EMAIL') and settings.ADMIN_EMAIL:
        destinatarios.append(settings.ADMIN_EMAIL)

    send_mail(
        subject=subject,
        message='',  # Se puede dejar vacío si se usa HTML
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=destinatarios,
        html_message=html_message,
        fail_silently=False, # Poner en True en producción
    )

def enviar_notificacion_nuevo_usuario(user, password_plano):
    """
    Función auxiliar para enviar un correo de bienvenida a un nuevo usuario.
    """
    if not user.email:
        return

    subject = "¡Bienvenido al Sistema de Gestión de Eventos!"
    contexto_email = {
        'user': user,
        'password': password_plano,
        'login_url': 'http://127.0.0.1:8000/login/', # Idealmente, obtener esto de forma dinámica
    }
    
    html_message = render_to_string('emails/notificacion_nuevo_usuario.html', contexto_email)
    
    send_mail(
        subject=subject,
        message='',
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        html_message=html_message,
        fail_silently=False,
    )

@login_required
def dashboard(request):
    """
    Vista para el dashboard principal. Muestra estadísticas y accesos directos.
    También ejecuta la actualización de estados de eventos de forma periódica.
    """
    # --- Lógica para actualización automática de estados ---
    # ADVERTENCIA: Esto se ejecuta en cada carga de página. Puede afectar el rendimiento.
    try:
        call_command('actualizar_estados_eventos')
    except Exception as e:
        # En caso de error, se podría registrar o notificar
        messages.error(request, f"Error al actualizar estados de eventos: {e}")

    # --- Lógica original del dashboard ---
    hoy = timezone.localdate()
    start_of_month = hoy.replace(day=1)
    next_month = hoy.replace(day=28) + timedelta(days=4)
    end_of_month = next_month - timedelta(days=next_month.day)

    # Base queryset for stats
    if request.user.is_admin_user():
        base_qs = Evento.objects.all()
    else:
        base_qs = Evento.objects.filter(usuario_creador=request.user)

    eventos_mes = base_qs.filter(fecha_hora_inicio__date__range=[start_of_month, end_of_month])
    
    total_eventos = eventos_mes.count()
    eventos_programados = eventos_mes.filter(estado='programado').count()
    eventos_finalizados = eventos_mes.filter(estado='finalizado').count()
    eventos_hoy_count = base_qs.filter(fecha_hora_inicio__date=hoy).count()

    # --- Filtering logic for event cards ---
    
    # Initial queryset for cards
    if request.user.is_admin_user():
        eventos_filtrados = Evento.objects.all()
    else:
        eventos_filtrados = Evento.objects.filter(usuario_creador=request.user)

    # Date period filter
    periodo = request.GET.get('periodo', 'dia')
    if periodo != 'todo':
        if periodo == 'dia':
            start_date = hoy
            end_date = hoy
        elif periodo == 'semana':
            start_date = hoy - timedelta(days=hoy.weekday())
            end_date = start_date + timedelta(days=6)
        else: # mes
            start_date = hoy.replace(day=1)
            end_date = end_of_month
        
        eventos_filtrados = eventos_filtrados.filter(fecha_hora_inicio__date__range=[start_date, end_date])

    # Apply filters from form
    form_filtro = FiltroEventosForm(request.GET)
    if form_filtro.is_valid():
        municipio = form_filtro.cleaned_data.get('municipio')
        if municipio:
            eventos_filtrados = eventos_filtrados.filter(municipio=municipio)

        dependencia = form_filtro.cleaned_data.get('dependencia')
        if dependencia:
            eventos_filtrados = eventos_filtrados.filter(usuario_creador__dependencia=dependencia)

        folio = form_filtro.cleaned_data.get('folio')
        if folio:
            eventos_filtrados = eventos_filtrados.filter(id=folio)

    eventos_filtrados = eventos_filtrados.order_by('fecha_hora_inicio')

    context = {
        'total_eventos': total_eventos,
        'eventos_programados': eventos_programados,
        'eventos_finalizados': eventos_finalizados,
        'eventos_hoy_count': eventos_hoy_count,
        'eventos_filtrados': eventos_filtrados,
        'form_filtro': form_filtro,
        'periodo': periodo,
        'titulo': 'Inicio'
    }
    
    return render(request, 'dashboard.html', context)

@login_required
@active_user_required
def exportar_eventos_excel(request):
    """Exportar eventos filtrados a Excel"""
    
    # Aplicar la misma lógica de filtros que en dashboard
    hoy = timezone.localdate()
    start_of_month = hoy.replace(day=1)
    next_month = hoy.replace(day=28) + timedelta(days=4)
    end_of_month = next_month - timedelta(days=next_month.day)

    if request.user.is_admin_user():
        eventos_filtrados = Evento.objects.all()
    else:
        eventos_filtrados = Evento.objects.filter(usuario_creador=request.user)

    # Filtro de período
    periodo = request.GET.get('periodo', 'dia')
    if periodo != 'todo':
        if periodo == 'dia':
            start_date = hoy
            end_date = hoy
        elif periodo == 'semana':
            start_date = hoy - timedelta(days=hoy.weekday())
            end_date = start_date + timedelta(days=6)
        else: # mes
            start_date = hoy.replace(day=1)
            end_date = end_of_month
        
        eventos_filtrados = eventos_filtrados.filter(fecha_hora_inicio__date__range=[start_date, end_date])

    # Aplicar filtros del formulario
    form_filtro = FiltroEventosForm(request.GET)
    if form_filtro.is_valid():
        municipio = form_filtro.cleaned_data.get('municipio')
        if municipio:
            eventos_filtrados = eventos_filtrados.filter(municipio=municipio)

        dependencia = form_filtro.cleaned_data.get('dependencia')
        if dependencia:
            eventos_filtrados = eventos_filtrados.filter(usuario_creador__dependencia=dependencia)

        folio = form_filtro.cleaned_data.get('folio')
        if folio:
            eventos_filtrados = eventos_filtrados.filter(id=folio)

    eventos_filtrados = eventos_filtrados.order_by('fecha_hora_inicio')

    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eventos"

    # Encabezados
    headers = ['Folio', 'Nombre del Evento', 'Fecha Inicio', 'Fecha Fin', 'Lugar', 'Municipio', 'Dependencia', 'Estado']
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="006554", end_color="006554", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Datos
    for row_num, evento in enumerate(eventos_filtrados, 2):
        ws.cell(row=row_num, column=1, value=evento.id)
        ws.cell(row=row_num, column=2, value=evento.nombre_evento)
        ws.cell(row=row_num, column=3, value=evento.fecha_hora_inicio.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=4, value=evento.fecha_hora_fin.strftime('%d/%m/%Y %H:%M') if evento.fecha_hora_fin else '')
        ws.cell(row=row_num, column=5, value=evento.lugar_evento or '')
        ws.cell(row=row_num, column=6, value=evento.municipio.nom_mun if evento.municipio else '')
        ws.cell(row=row_num, column=7, value=evento.dependencia or '')
        ws.cell(row=row_num, column=8, value=evento.get_estado_display())

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="eventos_{periodo}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
@active_user_required
@captura_or_admin_required
def crear_evento(request):
    """Vista para crear un nuevo evento"""
    
    if request.method == 'POST':
        form = EventoForm(request.POST, request.FILES, user=request.user)
        
        if form.is_valid():
            evento = form.save()
            
            # Enviar notificación por correo
            try:
                enviar_notificacion_evento(evento, subject=f"Nuevo evento: {evento.nombre_evento}", tipo_accion='creado')
                messages.success(
                    request, 
                    f'✅ Evento "{evento.nombre_evento}" creado exitosamente. '
                    f'Se ha enviado una notificación por correo electrónico.'
                )
            except Exception as e:
                messages.warning(
                    request,
                    f'✅ Evento "{evento.nombre_evento}" creado, pero hubo un error al enviar el correo: {e}'
                )
            
            return redirect('detalle_evento', evento_id=evento.id)
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        form = EventoForm(user=request.user)
    
    return render(request, 'eventos/crear_evento.html', {
        'form': form,
        'titulo': 'Nuevo Evento'
    })

@login_required
@active_user_required
@captura_or_admin_required
def mis_eventos(request):
    """Vista para listar eventos del usuario actual o todos (si es admin)"""
    
    # Inicializar queryset
    if request.user.is_admin_user():
        eventos = Evento.objects.all()
        titulo = 'Todos los Eventos del Sistema'
    else:
        eventos = request.user.eventos_creados.all()
        titulo = 'Mis Eventos'
    
    # Aplicar filtros
    form_filtro = FiltroEventosForm(request.GET)
    
    if form_filtro.is_valid():
        # Filtro por búsqueda de texto
        busqueda = form_filtro.cleaned_data.get('busqueda')
        if busqueda:
            eventos = eventos.filter(
                Q(nombre_evento__icontains=busqueda) |
                Q(lugar_evento__icontains=busqueda) |
                Q(numero_documento__icontains=busqueda) |
                Q(dependencia__icontains=busqueda)
            )
        
        # Filtro por estado
        estado = form_filtro.cleaned_data.get('estado')
        if estado:
            eventos = eventos.filter(estado=estado)
        
        # Filtro por asistencia del gobernador
        asistira_gobernador = form_filtro.cleaned_data.get('asistira_gobernador')
        if asistira_gobernador:
            eventos = eventos.filter(asistira_gobernador=asistira_gobernador == 'True')
        
        # Filtros por fecha
        fecha_desde = form_filtro.cleaned_data.get('fecha_desde')
        fecha_hasta = form_filtro.cleaned_data.get('fecha_hasta')
        
        if fecha_desde:
            eventos = eventos.filter(fecha_hora_inicio__date__gte=fecha_desde)
        
        if fecha_hasta:
            eventos = eventos.filter(fecha_hora_inicio__date__lte=fecha_hasta)
        
        # Filtro por municipio
        municipio = form_filtro.cleaned_data.get('municipio')
        if municipio:
            eventos = eventos.filter(municipio=municipio)
    
    # Ordenar por fecha de inicio (más recientes primero)
    eventos = eventos.order_by('-fecha_hora_inicio')
    
    # Paginación
    paginator = Paginator(eventos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas para mostrar
    total_eventos = eventos.count()
    eventos_programados = eventos.filter(estado='programado').count()
    eventos_activos = eventos.filter(estado='activo').count()
    eventos_finalizados = eventos.filter(estado='finalizado').count()
    
    return render(request, 'eventos/mis_eventos.html', {
        'page_obj': page_obj,
        'form_filtro': form_filtro,
        'titulo': titulo,
        'total_eventos': total_eventos,
        'eventos_programados': eventos_programados,
        'eventos_activos': eventos_activos,
        'eventos_finalizados': eventos_finalizados,
    })




@login_required
def calendar_view(request):
    """
    Vista para renderizar el calendario HTML del lado del servidor.
    """
    try:
        year = int(request.GET.get('year', timezone.now().year))
        month = int(request.GET.get('month', timezone.now().month))
    except ValueError:
        year = timezone.now().year
        month = timezone.now().month

    # Lógica de navegación de meses
    mes_siguiente = month + 1
    año_siguiente = year
    if mes_siguiente > 12:
        mes_siguiente = 1
        año_siguiente += 1

    mes_anterior = month - 1
    año_anterior = year
    if mes_anterior < 1:
        mes_anterior = 12
        año_anterior -= 1
    
    # Obtener eventos del mes
    if request.user.is_admin_user():
        eventos_mes = Evento.objects.filter(fecha_hora_inicio__year=year, fecha_hora_inicio__month=month)
    else:
        eventos_mes = Evento.objects.filter(
            usuario_creador=request.user,
            fecha_hora_inicio__year=year,
            fecha_hora_inicio__month=month
        )

    # Agrupar eventos por día
    eventos_por_dia = defaultdict(list)
    for evento in eventos_mes:
        local_time = timezone.localtime(evento.fecha_hora_inicio)
        eventos_por_dia[local_time.day].append(evento)

    # Serializar eventos para JavaScript
    eventos_para_json = defaultdict(list)
    for day, eventos_lista in eventos_por_dia.items():
        for evento in eventos_lista:
            local_start_time = timezone.localtime(evento.fecha_hora_inicio)
            local_end_time = timezone.localtime(evento.fecha_hora_fin) if evento.fecha_hora_fin else None

            eventos_para_json[str(day)].append({
                'id': evento.id,
                'title': evento.nombre_evento,
                'start': local_start_time.isoformat(),
                'end': local_end_time.isoformat() if local_end_time else None,
                'dependencia': evento.dependencia,
                'municipio': evento.municipio.nom_mun if evento.municipio else 'No especificado',
                'estado': evento.get_estado_display(),
            })

    # Generar el calendario
    cal = calendar.Calendar()
    calendario_semanas = cal.monthdayscalendar(year, month)
    
    # Nombre del mes en español
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    context = {
        'titulo': 'Calendario de Eventos',
        'calendario': calendario_semanas,
        'mes': month,
        'año': year,
        'mes_nombre': meses_es.get(month),
        'mes_siguiente': mes_siguiente,
        'año_siguiente': año_siguiente,
        'mes_anterior': mes_anterior,
        'año_anterior': año_anterior,
        'hoy': timezone.now().day if year == timezone.now().year and month == timezone.now().month else 0,
        'es_mes_actual': year == timezone.now().year and month == timezone.now().month,
        'eventos_por_dia': eventos_por_dia,
        'eventos_por_dia_json': json.dumps(eventos_para_json),
    }
    return render(request, 'eventos/calendario.html', context)


@login_required
@active_user_required
@captura_or_admin_required
def detalle_evento(request, evento_id):
    """Vista para ver detalles de un evento específico"""
    
    # Los usuarios de captura solo ven sus eventos, los admin ven todos
    if request.user.is_admin_user():
        evento = get_object_or_404(Evento, id=evento_id)
    else:
        evento = get_object_or_404(Evento, id=evento_id, usuario_creador=request.user)
    
    # Lógica para determinar si se puede editar
    if request.user.is_admin_user():
        puede_editar = True
    else: # Es usuario de captura
        puede_editar = evento.estado == 'programado' and request.user == evento.usuario_creador

    puede_cancelar = (
        request.user.is_admin_user() and 
        evento.estado in ['programado', 'revisado', 'activo']
    )
    
    return render(request, 'eventos/detalle_evento.html', {
        'evento': evento,
        'puede_editar': puede_editar,
        'puede_cancelar': puede_cancelar,
    })

@login_required
@active_user_required
@captura_or_admin_required
def editar_evento(request, evento_id):
    """Vista para editar un evento existente"""
    
    evento = get_object_or_404(Evento, id=evento_id)

    # Lógica de permisos de edición
    if request.user.is_admin_user():
        pass # El admin puede editar en cualquier estado
    else: # Usuario de captura
        if evento.usuario_creador != request.user:
            messages.error(request, 'No tienes permisos para editar este evento.')
            return redirect('dashboard')
        
        if evento.estado != 'programado':
            messages.error(request, f'No puedes editar un evento que ya ha sido "{evento.get_estado_display()}".')
            return redirect('detalle_evento', evento_id=evento.id)

    if request.method == 'POST':
        form = EventoForm(request.POST, request.FILES, instance=evento, user=request.user)
        
        if form.is_valid():
            evento_actualizado = form.save()
            
            # Enviar notificación por correo
            try:
                enviar_notificacion_evento(evento_actualizado, subject=f"Evento actualizado: {evento_actualizado.nombre_evento}", tipo_accion='actualizado')
                messages.success(
                    request,
                    f'✅ Evento "{evento_actualizado.nombre_evento}" actualizado correctamente. Se envió un correo.'
                )
            except Exception as e:
                messages.warning(
                    request,
                    f'✅ Evento "{evento_actualizado.nombre_evento}" actualizado, pero hubo un error al enviar el correo: {e}'
                )
            
            return redirect('detalle_evento', evento_id=evento_actualizado.id)
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        form = EventoForm(instance=evento, user=request.user)
    
    return render(request, 'eventos/crear_evento.html', {
        'form': form,
        'evento': evento,
        'titulo': f'Editar Evento: {evento.nombre_evento}',
        'editando': True
    })

@login_required
@admin_required
def cancelar_evento(request, evento_id):
    """Vista para cancelar un evento (solo admin)"""
    
    evento = get_object_or_404(Evento, id=evento_id)
    
    if evento.estado in ['finalizado', 'cancelado']:
        messages.error(request, 'No se puede cancelar un evento finalizado o ya cancelado.')
        return redirect('detalle_evento', evento_id=evento.id)
    
    if request.method == 'POST':
        estado_anterior = evento.estado
        evento.estado = 'cancelado'
        evento.save()
        
        # Crear log del cambio
        from .models import LogEventoEstado
        LogEventoEstado.objects.create(
            evento=evento,
            estado_anterior=estado_anterior,
            estado_nuevo='cancelado',
            usuario_cambio=request.user,
            comentario='Evento cancelado manualmente por administrador',
            automatico=False
        )
        
        messages.success(
            request,
            f'✅ Evento "{evento.nombre_evento}" cancelado correctamente. '
            f'Se enviará notificación por correo.'
        )
        
        return redirect('detalle_evento', evento_id=evento.id)
    
    return render(request, 'eventos/confirmar_cancelacion.html', {'evento': evento})

@login_required
@active_user_required
@captura_or_admin_required
def descargar_pdf(request, evento_id):
    """Vista para descargar el PDF de un evento"""
    
    # Verificar permisos
    if request.user.is_admin_user():
        evento = get_object_or_404(Evento, id=evento_id)
    else:
        evento = get_object_or_404(Evento, id=evento_id, usuario_creador=request.user)
    
    if not evento.archivo_pdf:
        messages.error(request, 'Este evento no tiene archivo PDF asociado.')
        return redirect('detalle_evento', evento_id=evento.id)
    
    # Verificar que el archivo existe
    if not os.path.exists(evento.archivo_pdf.path):
        messages.error(request, 'El archivo PDF no se encuentra en el servidor.')
        return redirect('detalle_evento', evento_id=evento.id)
    
    try:
        with open(evento.archivo_pdf.path, 'rb') as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{evento.numero_documento}.pdf"'
            return response
    except Exception as e:
        messages.error(request, f'Error al descargar el archivo: {str(e)}')
        return redirect('detalle_evento', evento_id=evento.id)

@login_required
def perfil_usuario(request):
    """
    Vista para que el usuario vea y edite su perfil.
    """
    user = request.user
    profile_form = PerfilUsuarioForm(instance=user)
    password_form = CustomPasswordChangeForm(user)

    if request.method == 'POST':
        # Distinguir qué formulario se envió
        if 'update_profile' in request.POST:
            profile_form = PerfilUsuarioForm(request.POST, instance=user)
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, '✅ Tu perfil ha sido actualizado correctamente.')
                return redirect('perfil_usuario')
            else:
                messages.error(request, '❌ Por favor, corrige los errores al actualizar tu perfil.')

        elif 'change_password' in request.POST:
            password_form = CustomPasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                # Actualizar la sesión del usuario para que no se cierre
                update_session_auth_hash(request, user)
                messages.success(request, '🔑 Tu contraseña ha sido cambiada exitosamente.')
                return redirect('perfil_usuario')
            else:
                messages.error(request, '❌ Por favor, corrige los errores al cambiar tu contraseña.')

    context = {
        'profile_form': profile_form,
        'password_form': password_form,
        'titulo': 'Mi Perfil',
        'user': user
    }
    return render(request, 'perfil_usuario.html', context)


@login_required
@admin_required
def crear_usuario(request):
    """
    Vista para que un administrador cree un nuevo usuario de captura.
    """
    if request.method == 'POST':
        form = AdminCrearUsuarioForm(request.POST)
        if form.is_valid():
            password_plano = form.cleaned_data.get('password')
            user = form.save()
            
            try:
                enviar_notificacion_nuevo_usuario(user, password_plano)
                messages.success(request, f'✅ Usuario "{user.username}" creado exitosamente. Se ha enviado un correo de bienvenida.')
            except Exception as e:
                messages.warning(request, f'✅ Usuario "{user.username}" creado, pero hubo un error al enviar el correo de bienvenida: {e}')

            # Redirigir a la lista de usuarios para ver el resultado.
            return redirect('gestionar_usuarios')

        else:
            messages.error(request, '❌ Por favor, corrige los errores en el formulario.')
    else:
        form = AdminCrearUsuarioForm()

    context = {
        'form': form,
        'titulo': 'Crear Nuevo Usuario de Captura'
    }
    return render(request, 'usuarios/crear_usuario.html', context)

@login_required
@admin_required
def editar_usuario(request, user_id):
    """
    Vista para que un administrador edite un usuario existente.
    """
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        form = AdminEditarUsuarioForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Usuario "{user.username}" actualizado exitosamente.')
            return redirect('gestionar_usuarios')
        else:
            messages.error(request, '❌ Por favor, corrige los errores en el formulario.')
    else:
        form = AdminEditarUsuarioForm(instance=user)

    context = {
        'form': form,
        'titulo': f'Editar Usuario: {user.username}',
        'usuario_editado': user
    }
    return render(request, 'usuarios/editar_usuario.html', context)

@login_required
@admin_required
def gestionar_usuarios(request):
    """
    Vista para que un administrador vea, filtre y gestione usuarios.
    """
    # Excluir al superusuario de la lista para evitar que se desactive a sí mismo
    usuarios_list = CustomUser.objects.filter(is_superuser=False).order_by('username')
    dependencias_list = Dependencia.objects.all()

    # Lógica de búsqueda y filtro
    query = request.GET.get('q', '')
    dependencia_id = request.GET.get('dependencia', '')

    if query:
        usuarios_list = usuarios_list.filter(
            Q(username__icontains=query) |
            Q(nombre_completo__icontains=query) |
            Q(email__icontains=query)
        )

    if dependencia_id:
        usuarios_list = usuarios_list.filter(dependencia_id=dependencia_id)

    context = {
        'usuarios': usuarios_list,
        'titulo': 'Gestionar Usuarios',
        'query': query,
        'dependencias_list': dependencias_list,
        'dependencia_id': dependencia_id,
    }
    return render(request, 'usuarios/gestionar_usuarios.html', context)

@login_required
@admin_required
def toggle_user_status(request, user_id):
    return HttpResponse(f"Toggle user status for user {user_id}")

@login_required
def check_user_permissions(request):
    return HttpResponse("Check user permissions")

@login_required
@admin_required
def lista_municipios(request):
    """
    Vista para listar todos los municipios.
    """
    municipios = Municipio.objects.all()
    
    context = {
        'municipios': municipios,
        'titulo': 'Lista de Municipios'
    }
    
    return render(request, 'eventos/lista_municipios.html', context)